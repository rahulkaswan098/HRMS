import json
import os
from typing import Dict, List
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime

class OutputGenerator:
    """Generate output files (Excel, JSON) from screening results"""
    
    def __init__(self, output_folder: str = 'outputs'):
        self.output_folder = output_folder
        if not os.path.exists(output_folder):
            os.makedirs(output_folder)
    
    def generate_excel(self, results: Dict, session_id: str) -> str:
        """
        Generate Excel report from screening results
        
        Args:
            results: Screening results dictionary
            session_id: Session identifier
            
        Returns:
            Path to generated Excel file
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Screening Results"
        
        # Headers
        headers = ['Rank', 'Name', 'Match Score', 'Experience (Years)', 
                  'Current Role', 'Current Company', 'Education', 
                  'Recommendation', 'Email', 'Phone']
        
        # Style headers
        header_fill = PatternFill(start_color="E50914", end_color="E50914", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Data rows
        candidates = results.get('candidates', [])
        for idx, candidate in enumerate(candidates, 1):
            row = [
                idx,
                candidate.get('name', 'Unknown'),
                candidate.get('match_score', 0),
                candidate.get('experience_years', 0),
                candidate.get('current_role', 'N/A'),
                candidate.get('current_company', 'N/A'),
                candidate.get('education', 'N/A'),
                candidate.get('recommendation', 'N/A'),
                candidate.get('email', 'N/A'),
                candidate.get('phone', 'N/A')
            ]
            
            for col, value in enumerate(row, 1):
                ws.cell(row=idx+1, column=col, value=value)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save file
        filename = f"screening_results_{session_id}.xlsx"
        filepath = os.path.join(self.output_folder, filename)
        wb.save(filepath)
        
        return filepath
    
    def generate_json(self, results: Dict, session_id: str) -> str:
        """
        Generate JSON file from screening results
        
        Args:
            results: Screening results dictionary
            session_id: Session identifier
            
        Returns:
            Path to generated JSON file
        """
        # Add metadata
        output = {
            'session_id': session_id,
            'generated_at': datetime.now().isoformat(),
            'results': results
        }
        
        filename = f"screening_results_{session_id}.json"
        filepath = os.path.join(self.output_folder, filename)
        
        with open(filepath, 'w', encoding='utf-8') as f:
            json.dump(output, f, indent=2, ensure_ascii=False)
        
        return filepath
    
    def generate_candidate_report(self, candidate: Dict, session_id: str) -> str:
        """
        Generate detailed HTML report for a single candidate
        
        Args:
            candidate: Candidate dictionary
            session_id: Session identifier
            
        Returns:
            HTML string
        """
        html = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <title>Candidate Report - {candidate.get('name', 'Unknown')}</title>
            <style>
                body {{
                    font-family: Arial, sans-serif;
                    max-width: 800px;
                    margin: 0 auto;
                    padding: 20px;
                    background: #f5f5f5;
                }}
                .header {{
                    background: #e50914;
                    color: white;
                    padding: 20px;
                    border-radius: 8px;
                    margin-bottom: 20px;
                }}
                .section {{
                    background: white;
                    padding: 20px;
                    margin-bottom: 15px;
                    border-radius: 8px;
                    box-shadow: 0 2px 4px rgba(0,0,0,0.1);
                }}
                .score {{
                    font-size: 48px;
                    font-weight: bold;
                    color: #e50914;
                }}
                .label {{
                    font-weight: bold;
                    color: #666;
                }}
                ul {{
                    list-style-type: none;
                    padding-left: 0;
                }}
                li {{
                    padding: 5px 0;
                    border-bottom: 1px solid #eee;
                }}
            </style>
        </head>
        <body>
            <div class="header">
                <h1>{candidate.get('name', 'Unknown')}</h1>
                <p>{candidate.get('current_role', 'N/A')} at {candidate.get('current_company', 'N/A')}</p>
            </div>
            
            <div class="section">
                <div class="score">{candidate.get('match_score', 0)}%</div>
                <p>Match Score</p>
                <p><strong>Recommendation:</strong> {candidate.get('recommendation', 'N/A')}</p>
            </div>
            
            <div class="section">
                <h2>Contact Information</h2>
                <p><span class="label">Email:</span> {candidate.get('email', 'N/A')}</p>
                <p><span class="label">Phone:</span> {candidate.get('phone', 'N/A')}</p>
            </div>
            
            <div class="section">
                <h2>Experience</h2>
                <p><span class="label">Years:</span> {candidate.get('experience_years', 0)}</p>
                <p><span class="label">Current Role:</span> {candidate.get('current_role', 'N/A')}</p>
                <p><span class="label">Current Company:</span> {candidate.get('current_company', 'N/A')}</p>
            </div>
            
            <div class="section">
                <h2>Education</h2>
                <p>{candidate.get('education', 'N/A')}</p>
            </div>
            
            <div class="section">
                <h2>Key Skills</h2>
                <ul>
                    {''.join([f'<li>{skill}</li>' for skill in candidate.get('skills', [])])}
                </ul>
            </div>
            
            <div class="section">
                <h2>Strengths</h2>
                <ul>
                    {''.join([f'<li>✓ {strength}</li>' for strength in candidate.get('strengths', [])])}
                </ul>
            </div>
            
            <div class="section">
                <h2>Concerns</h2>
                <ul>
                    {''.join([f'<li>⚠ {concern}</li>' for concern in candidate.get('concerns', [])])}
                </ul>
            </div>
            
            <div class="section">
                <h2>Summary</h2>
                <p>{candidate.get('summary', 'No summary available')}</p>
            </div>
        </body>
        </html>
        """
        
        return html
